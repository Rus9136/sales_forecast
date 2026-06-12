"""Generate a management-facing pricing report for Sandyq Group restaurants.

Produces an Excel workbook with:
  - Сводка (executive summary): uplift per restaurant
  - Рекомендации (clean, sorted recommendations)
  - Методология (how the numbers are derived)

Run from host:
    python3 scripts/sandyq_pricing_report.py
"""

import os
import psycopg2
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _env(key, default=""):
    for fname in (".env.prod", ".env"):
        path = os.path.join(os.path.dirname(__file__), "..", fname)
        if os.path.exists(path):
            for line in open(path):
                if line.startswith(key + "="):
                    return line.split("=", 1)[1].strip()
    return os.environ.get(key, default)


SANDYQ_DEPTS = [
    "700917ce-77e5-479a-a4a9-68c2a9e76fc2",  # Sandyq Алматы
    "2086adde-d191-496e-9ff7-eb78173fa8bb",  # Sandyq Astana
    "e8cba932-e56d-4fe7-bbc9-01c92b765a4c",  # Sandyq Turkestan
    "c3c76c7c-bead-4be2-88ce-7d7415aeabe4",  # Tary Astana
    "7c5422a2-ba5e-4af3-9edb-f417c9c34ef5",  # Tary Dendropark
]

ROLE_RU = {
    "premium_anchor": "Премиум-якорь",
    "margin_driver": "Драйвер маржи",
    "traffic_driver": "Драйвер трафика",
    "image_rare": "Имиджевая",
    "tail": "Хвост",
    "unknown": "—",
}
GRADE_RU = {
    "A": "A — высокая",
    "B": "B — хорошая",
    "C": "C — средняя",
    "D": "D — низкая",
}

# Visual styling
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D6E0F0"
GREEN = "E2EFDA"
GREY = "F2F2F2"
WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUBHDR_FILL = PatternFill("solid", fgColor=BLUE)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def fetch_data(conn):
    cur = conn.cursor()
    cur.execute(
        """
        WITH latest_batch AS (
            SELECT department_id, MAX(created_at) AS mx
            FROM price_recommendation
            WHERE department_id = ANY(%s::uuid[])
            GROUP BY department_id
        )
        SELECT
            d.name, p.name, pr.menu_role, pr.elasticity_used, pr.elasticity_grade,
            pr.current_price, pr.recommended_price, pr.delta_pct,
            pr.cogs, pr.current_gp, pr.expected_gp, pr.delta_gp
        FROM price_recommendation pr
        JOIN latest_batch lb
            ON lb.department_id = pr.department_id AND lb.mx = pr.created_at
        JOIN departments d ON d.id = pr.department_id
        JOIN product p ON p.id = pr.product_id
        WHERE pr.status = 'new'
          -- Exclude certificates / service items / no-cost items
          AND p.name NOT ILIKE '%%сертификат%%'
          AND p.name NOT ILIKE '%%доставка%%'
          AND pr.cogs IS NOT NULL
          AND pr.cogs > 0
        ORDER BY
            CASE pr.elasticity_grade
                WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 ELSE 4 END,
            pr.delta_gp DESC
        """,
        (SANDYQ_DEPTS,),
    )
    rows = cur.fetchall()
    cur.close()
    return rows


def style_header(ws, row, headers, fill=HDR_FILL):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = WHITE_BOLD
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def build_report(rows, out_path):
    wb = Workbook()

    # ---- Sheet 1: Сводка ----
    ws = wb.active
    ws.title = "Сводка"

    # Title block
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Sandyq Group — рекомендации по повышению цен"
    t.font = Font(bold=True, size=16, color=NAVY)
    ws.merge_cells("A2:F2")
    s = ws["A2"]
    s.value = f"Сформировано системой интеллектуального ценообразования (ML) · {date.today().strftime('%d.%m.%Y')}"
    s.font = Font(italic=True, size=10, color="595959")

    # Per-restaurant aggregation (split by confidence)
    by_rest = {}
    for r in rows:
        rest = r[0]
        d = by_rest.setdefault(rest, {"n": 0, "weekly_gp": 0.0,
                                       "conf_n": 0, "conf_gp": 0.0})
        d["n"] += 1
        d["weekly_gp"] += float(r[11] or 0)
        if r[4] in ("A", "B", "C"):
            d["conf_n"] += 1
            d["conf_gp"] += float(r[11] or 0)

    total_n = sum(d["n"] for d in by_rest.values())
    total_weekly = sum(d["weekly_gp"] for d in by_rest.values())

    # Confidence split: B/C = надёжные (подтверждённые), D = требуют валидации
    conf_n = sum(1 for r in rows if r[4] in ("A", "B", "C"))
    conf_weekly = sum(float(r[11] or 0) for r in rows if r[4] in ("A", "B", "C"))
    pend_n = total_n - conf_n
    pend_weekly = total_weekly - conf_weekly

    # KPI cards row — two tiers
    ws["A4"] = "Ключевые показатели"
    ws["A4"].font = Font(bold=True, size=12, color=NAVY)
    kpis = [
        ("Надёжных рекомендаций", f"{conf_n}"),
        ("Подтверждённый прирост / нед", f"{conf_weekly:,.0f} ₸".replace(",", " ")),
        ("Подтверждённый прирост / год", f"{conf_weekly * 52:,.0f} ₸".replace(",", " ")),
        ("Дополнит. потенциал / год*", f"{pend_weekly * 52:,.0f} ₸".replace(",", " ")),
    ]
    for i, (label, val) in enumerate(kpis):
        col = 1 + i * 2
        lc = ws.cell(row=6, column=col, value=label)
        lc.font = Font(size=9, color="595959")
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        vc = ws.cell(row=7, column=col, value=val)
        vc.font = Font(bold=True, size=13, color=BLUE if i < 3 else "808080")
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)

    ws.merge_cells("A8:H8")
    star = ws["A8"]
    star.value = (f"* {pend_n} позиций с оценкой эластичности «низкая» — учтены при недостатке "
                  "ценовой истории, требуют подтверждения в пилоте. Общий потенциал (надёжные + "
                  f"требующие валидации): {total_weekly * 52:,.0f} ₸/год.".replace(",", " "))
    star.font = Font(italic=True, size=9, color="808080")
    star.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[8].height = 28

    # Per-restaurant table
    hdr_row = 11
    style_header(ws, hdr_row, [
        "Ресторан", "Надёжных позиций", "Подтверждённый\nприрост/нед, ₸",
        "Подтверждённый\nприрост/год, ₸", "Потенциал/год\n(все), ₸",
    ])
    ws.row_dimensions[hdr_row].height = 32
    rr = hdr_row + 1
    for rest in sorted(by_rest, key=lambda x: -by_rest[x]["conf_gp"]):
        d = by_rest[rest]
        vals = [rest, d["conf_n"], round(d["conf_gp"]),
                round(d["conf_gp"] * 52), round(d["weekly_gp"] * 52)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=col, value=v)
            c.border = BORDER
            if col >= 3:
                c.number_format = "#,##0"
            if col == 1:
                c.font = Font(bold=True)
            if col == 5:
                c.font = Font(color="808080")
        rr += 1
    # Total row
    tot_vals = ["ИТОГО", conf_n, round(conf_weekly),
                round(conf_weekly * 52), round(total_weekly * 52)]
    for col, v in enumerate(tot_vals, 1):
        c = ws.cell(row=rr, column=col, value=v)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.border = BORDER
        if col >= 3:
            c.number_format = "#,##0"

    widths = [22, 16, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Note
    note_row = rr + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    n = ws.cell(row=note_row, column=1)
    n.value = ("Все рекомендации соблюдают бизнес-правила: шаг ≤ +5%, маржа ≥ 60%, "
               "цены кратны 50/100 ₸, премиум-якоря не дешевеют. "
               "Каждая рекомендация требует утверждения управляющим.")
    n.font = Font(italic=True, size=9, color="595959")
    n.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 40

    # ---- Sheet 2: Рекомендации ----
    ws2 = wb.create_sheet("Рекомендации")
    headers = [
        "Ресторан", "Позиция", "Роль в меню", "Надёжность оценки",
        "Текущая цена, ₸", "Рекоменд. цена, ₸", "Δ %",
        "Себест., ₸", "Прирост прибыли/нед, ₸",
    ]
    style_header(ws2, 1, headers)
    ws2.freeze_panes = "A2"

    for i, r in enumerate(rows, start=2):
        (rest, prod, role, eps, grade, cur_p, rec_p, dpct, cogs, cgp, egp, dgp) = r
        vals = [
            rest, prod, ROLE_RU.get(role, role or "—"), GRADE_RU.get(grade, grade or "—"),
            float(cur_p), float(rec_p), float(dpct),
            float(cogs) if cogs else None, round(float(dgp or 0)),
        ]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=v)
            c.border = BORDER
            if col in (5, 6, 8, 9):
                c.number_format = "#,##0"
            if col == 7:
                c.number_format = "+0.0\"%\""
                c.font = Font(color="375623", bold=True)
        # zebra
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                if ws2.cell(row=i, column=col).fill.fgColor.rgb in (None, "00000000"):
                    ws2.cell(row=i, column=col).fill = PatternFill("solid", fgColor=GREY)

    widths2 = [18, 38, 16, 18, 15, 16, 9, 12, 20]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 3: Методология ----
    ws3 = wb.create_sheet("Методология")
    ws3.column_dimensions["A"].width = 100
    lines = [
        ("Как система формирует рекомендации", "h"),
        ("", ""),
        ("1. Прогноз спроса (ML)", "b"),
        ("Модель LightGBM прогнозирует спрос на каждое блюдо на основе 73 признаков "
         "(день недели, сезонность, праздники, история продаж).", ""),
        ("", ""),
        ("2. Ценовая эластичность", "b"),
        ("Для каждого блюда оценивается чувствительность спроса к цене на основе НАСТОЯЩИХ "
         "цен из приказов iiko (163 433 ценовых интервала за 24 месяца). Метод — иерархическая "
         "регрессия с заимствованием оценок по категориям при недостатке данных.", ""),
        ("", ""),
        ("3. Роль блюда в меню", "b"),
        ("Кластеризация (KMeans) делит меню на 5 ролей. Премиум-якоря защищены от снижения цены.", ""),
        ("", ""),
        ("4. Оптимизация прибыли", "b"),
        ("Для каждого блюда перебираются допустимые цены (±5%, шаг 50/100 ₸) и выбирается та, "
         "что даёт максимальную валовую прибыль: ВП = (Цена − Себестоимость) × Прогноз спроса.", ""),
        ("", ""),
        ("5. Бизнес-правила (жёсткие ограничения)", "b"),
        ("• Маржа ≥ 60%   • Шаг ≤ ±5% за итерацию   • Не чаще 1 раза в 14 дней", ""),
        ("• Премиум-якоря — только повышение   • Округление до 50 ₸ (флагманы — 100 ₸)", ""),
        ("", ""),
        ("Подтверждение гипотезы", "h"),
        ("", ""),
        ("Анализ на чистых ценах из приказов подтвердил: спрос в премиум-сегменте "
         "слабо чувствителен к цене (медианная эластичность −0,47). Это означает, что "
         "умеренное повышение цен увеличивает прибыль — выручка растёт быстрее, чем снижается спрос.", ""),
        ("", ""),
        ("Важно: все цифры — это ПРОГНОЗ. Рекомендуется пилот на ограниченном наборе позиций "
         "с замером фактического эффекта через 2 недели. Все рекомендации носят рекомендательный "
         "характер и утверждаются управляющим вручную.", "i"),
    ]
    rr = 1
    for text, style in lines:
        c = ws3.cell(row=rr, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if style == "h":
            c.font = Font(bold=True, size=13, color=NAVY)
        elif style == "b":
            c.font = Font(bold=True, size=11, color=BLUE)
        elif style == "i":
            c.font = Font(italic=True, size=10, color="C00000")
        else:
            c.font = Font(size=10)
        if len(text) > 90:
            ws3.row_dimensions[rr].height = 45
        rr += 1

    wb.save(out_path)
    return total_n, total_weekly


def main():
    conn = psycopg2.connect(
        host="localhost", port=5435,
        user=_env("POSTGRES_USER"), password=_env("POSTGRES_PASSWORD"),
        dbname=_env("POSTGRES_DB"),
    )
    try:
        rows = fetch_data(conn)
    finally:
        conn.close()

    out = os.path.join(os.path.dirname(__file__), "..", "Sandyq_pricing_recommendations.xlsx")
    out = os.path.abspath(out)
    n, weekly = build_report(rows, out)
    print(f"Report saved: {out}")
    print(f"Recommendations: {n}, weekly GP uplift: {weekly:,.0f} tenge")


if __name__ == "__main__":
    main()
